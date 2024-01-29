from tabulate import tabulate
from openpyxl import Workbook


# Main function for Generating the Travel List with all the arguments
def generate_travel_list(destination, activities, components, places, duration, weather, people):
    travel_list = {
        "Destination": destination,
        "Places": places,
        "Duration": duration,
        "Activities": activities,
        "Components": components,
        "Weather": weather,
        "People": people
    }
    return travel_list


# Function to Save the generated list as an Excel File by giving name only without extension
def save_to_excel(travel_list, filename):
    # Ensure the filename has the .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    wb = Workbook()
    ws = wb.active

    # Write headers
    ws['A1'] = "Destination"
    ws['B1'] = "Place"
    ws['C1'] = "Duration"
    ws['D1'] = "Activities"
    ws['E1'] = "Components"
    ws['F1'] = "Weather"
    ws['G1'] = "People"

    # Write data
    ws['A2'] = travel_list['Destination']
    ws['B2'] = "\n".join(travel_list['Places'])
    ws['C2'] = travel_list['Duration']
    ws['D2'] = "\n".join(travel_list['Activities'])
    ws['E2'] = "\n".join(travel_list['Components'])
    ws['F2'] = travel_list['Weather']
    ws['G2'] = travel_list['People']

    wb.save(filename)
    print(f"Travel list saved to {filename}")


# Display travel list function show complete list and all options added with Category and Details
def display_travel_list(travel_list):
    print("\nGenerated Travel List:")
    table = []

    for key, value in travel_list.items():
        if key == "Activities" or key == "Components" or key == "Places":
            table.append([key, "\n".join(value)])
        else:
            table.append([key, value])

    print(tabulate(table, headers=["Category", "Details"], tablefmt="grid"))


# Add activity is called for adding the activities by user
def add_activity(travel_list):
    new_activity = input("Enter the activity to add: ")
    travel_list["Activities"].append(new_activity)
    print(f"Added '{new_activity}' to the travel list.")


# Removing the necessary components like if mistakenly added
def remove_necessary_component(travel_list):
    display_travel_list(travel_list)
    component_to_remove = input("Enter the Necessary item to remove: ")
    if component_to_remove in travel_list["Components"]:
        travel_list["Components"].remove(component_to_remove)
        print(f"Removed '{component_to_remove}' from the travel list.")
    else:
        print(f"'{component_to_remove}' not found in the travel list.")


# Funtion to addd the necessary component item to carry with on the trip
def add_necessary_components(travel_list):
    while True:
        new_component = input("Enter a necessary component (or 'C' to complete): ")

        if new_component.upper() == 'C':
            break

        travel_list["Components"].append(new_component)
        print(f"Added '{new_component}' to the travel list.")


# Function to select the place of stay from predefined option in travel list generator
def choose_place(travel_list):
    print("\nSelect the type of place:")
    print("1. Hotel")
    print("2. Hostel")
    print("3. Short-term rental")
    print("4. Friends' house")
    print("5. Camping")

    choice = input("Enter your choice (1/2/3/4/5): ")

    places_dict = {
        '1': 'Hotel',
        '2': 'Hostel',
        '3': 'Short-term rental',
        '4': 'Friends\' house',
        '5': 'Camping'
    }

    if choice in places_dict:
        travel_list["Places"].append(places_dict[choice])
        print(f"Added '{places_dict[choice]}' to the travel list.")
    else:
        print("Invalid choice. Defaulting to 'Unknown'")
        travel_list["Places"].append("Unknown")


# Function to select the people those going on trip from predefined option in travel list generator
def choose_people(travel_list):
    print("\nSelect the type of place:")
    print("1. Male")
    print("2. Female")
    print("3. Baby")
    print("4. Both Male and Female")
    print("5. All of Husband, Wife and Child")

    choice = input("Enter your choice (1/2/3/4/5): ")

    people_dict = {
        '1': 'Male',
        '2': 'Female',
        '3': 'Baby',
        '4': 'Both\' Male\' &\' Female',
        '5': 'All\' of\' these\' Husband,\' Wife\' &\' Child'
    }

    if choice in people_dict:
        travel_list["People"] = people_dict[choice]
        print(f"Added '{people_dict[choice]}' to the travel list.")
    else:
        print("Invalid choice. Defaulting to 'Unknown'")
        travel_list["People"].append("Unknown")


# Function to add the duration of the trip like in days or months it will of months or days.
def choose_duration(travel_list):
    print("\nSelect the duration of your stay:")
    print("1. Number of days")
    print("2. Number of months")

    choice = input("Enter your choice (1/2): ")

    duration_dict = {
        '1': 'Number of days',
        '2': 'Number of months'
    }

    if choice in duration_dict:
        travel_list["Duration"] = duration_dict[choice]

        if choice == '1':
            days = input("Enter the number of days: ")
            travel_list["Duration"] += f" ({days} days)"
        elif choice == '2':
            months = input("Enter the number of months: ")
            travel_list["Duration"] += f" ({months} months)"

        print(f"Added '{travel_list['Duration']}' to the travel list.")
    else:
        print("Invalid choice. Defaulting to 'Unknown'")
        travel_list["Duration"] = "Unknown"


# Function to weather from suggested option in travel list generator
def choose_weather(travel_list):
    print("\nSelect the expected weather:")
    print("1. Freezing Cold")
    print("2. Cool")
    print("3. Warm")
    print("4. Hot")
    print("5. Rainy")
    print("6. Cloudy")

    choice = input("Enter your choice (1/2/3/4/5/6): ")

    weather_dict = {
        '1': 'Freezing Cold',
        '2': 'Cool',
        '3': 'Warm',
        '4': 'Hot',
        '5': 'Rainy',
        '6': 'Cloudy'
    }

    if choice in weather_dict:
        travel_list["Weather"] = weather_dict[choice]
        print(f"Added '{weather_dict[choice]}' to the travel list.")
    else:
        print("Invalid choice. Defaulting to 'Unknown'")
        travel_list["Weather"] = "Unknown"


# Function Defined to get destination of trip in same country city or in different country
def get_destination():
    print("\nSelect your destination:")
    print("1. Different country")
    print("2. Different city in my country")
    choice = input("Enter your choice (1/2): ")

    if choice == '1':
        return input("Enter the name of the country: ")
    elif choice == '2':
        return input("Enter the name of the city in your country: ")
    else:
        print("Invalid choice. Defaulting to 'Unknown'")
        return "Unknown"


# Main function to Call travel list function and also menu options are defined within it
def main():
    print("Welcome to the Travel List Generator!")

    destination = get_destination()
    activities = []
    components = []
    places = []
    duration = ""
    weather = ""
    people = ""

    travel_list = generate_travel_list(destination, activities, components, places, duration, weather, people)

    while True:
        print("\nMenu Options:")
        print("1. Display Travel List")
        print("2. Add Activity to the List")
        print("3. Remove Activity from the List")
        print("4. Add Necessary Item")
        print("5. Choose Place Type")
        print("6. Choose Duration of Stay")
        print("7. Choose Expected Weather")
        print("8. Save as Excel and Exit")
        print("9. Chose Gender for Trip")
        print("10. Exit Travel List Generator")

        choice = input("Enter your choice (1/2/3/4/5/6/7/8/9/10): ")

        if choice == '1':
            display_travel_list(travel_list)
        elif choice == '2':
            add_activity(travel_list)
        elif choice == '3':
            remove_necessary_component(travel_list)
        elif choice == '4':
            add_necessary_components(travel_list)
        elif choice == '5':
            choose_place(travel_list)
        elif choice == '6':
            choose_duration(travel_list)
        elif choice == '7':
            choose_weather(travel_list)
        elif choice == '8':
            excel_filename = input("Enter the Excel filename to save (without extension): ")
            save_to_excel(travel_list, excel_filename)
            print("Exiting Travel List Generator. Your travel list has been saved.")
            break
        elif choice == '9':
            choose_people(travel_list)
        elif choice == '10':
            print("Exiting Travel List Generator.")
            break


if __name__ == "__main__":
    main()
